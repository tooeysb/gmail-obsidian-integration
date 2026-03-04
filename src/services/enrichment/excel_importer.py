"""
Excel importer for CRM enrichment data.
Parses multi-tab spreadsheets with per-tab column normalization.
"""

import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from src.core.logging import get_logger

logger = get_logger(__name__)

# Mapping of normalized header keywords to canonical field names, per tab.
# Each tab has a dict of {canonical_key: [possible_header_substrings]}.
# Order matters: first match wins, so put more specific patterns first.
_TAB_COLUMN_MAPS: dict[str, dict[str, list[str]]] = {
    "Over 1M Customers": {
        "company_name": ["account name"],
        "arr": ["annual access value"],
        "revenue_segment": ["revenue segment"],
        "company_type": ["company type"],
        "billing_state": ["billing state/province"],
        "account_owner": ["account owner"],
        "renewal_date": ["next renewal start date"],
        "type": ["type"],  # "Customer - Annual" etc.
    },
    "Customer": {
        "name": ["name"],
        "company": ["company"],
        "title": ["title"],
        "affiliation_notes": ["affiliation"],
        "email": ["email address"],
        "personal_email": ["personal email"],
        "phone": ["phone call"],
        "ae": ["ae"],
        "csm": ["csm"],
    },
    "CAB-rolled in": {
        "first_name": ["first name"],
        "last_name": ["last name"],
        "title": ["title"],
        "company": ["company"],
        "salesforce_id": ["salesforce id"],
        "cab_type": ["cab"],
        "year": ["year"],
        "active_alumni": ["active or alumni"],
        "email": ["email"],
        "address": ["address"],
        "csm": ["csm"],
        "ae": ["ae"],
    },
    "2025 CAB": {
        "first_name": ["first name"],
        "last_name": ["last name"],
        "title": ["title"],
        "company": ["company"],
        "email": ["email address"],
        "notes": ["notes"],
    },
    "Cab List": {
        "location": ["location"],
        "first_name": ["first name"],
        "last_name": ["last name"],
        "title": ["title"],
        "company": ["company"],
        "salesforce_id": ["salesforce id"],
        "ecab": ["ecab"],
        "year": ["year"],
        "active_alumni": ["active or alumni"],
        "email": ["email"],
        "address": ["address"],
    },
    "Feb 7 Transition": {
        "arr": ["arr"],
        "renewal_date": ["renewal"],
        "name": ["name"],
        "company": ["company"],
        "title": ["title"],
        "affiliation_notes": ["affiliation"],
        "email": ["email address"],
        "personal_email": ["personal email"],
        "phone": ["phone call"],
        "ae": ["ae"],
        "csm": ["csm"],
    },
    "Margaux - Responses": {
        "name": ["name"],
        "company": ["company"],
        "personal_email": ["personal email"],
        "email": ["email"],
        "ae": ["account executive"],
        "csm": ["customer success"],
    },
    "Family_Friends_Mentors_Others": {
        # Column 0 has names but no header — handled as special case
        "email": ["email"],
        "phone": ["phone"],
        "notes": ["notes"],
    },
    "Internal": {
        "name": ["name"],
        "cell": ["cell"],
        "slack": ["slack"],
        "tlt1_status": ["tlt - 1", "tlt-1"],
    },
}

# Exact sheet name matching to avoid false positives like "Customer Sources" -> "Customer"
_EXACT_SHEET_NAMES: dict[str, str] = {
    "over 1m customers": "Over 1M Customers",
    "customer": "Customer",
    "cab-rolled in": "CAB-rolled in",
    "2025 cab": "2025 CAB",
    "cab list": "Cab List",
    "feb 7 transition": "Feb 7 Transition",
    "margaux - responses": "Margaux - Responses",
    "family_friends_mentors_others": "Family_Friends_Mentors_Others",
    "internal": "Internal",
}


def _extract_hyperlink_text(value: str) -> str:
    """Extract display text from Excel HYPERLINK formula strings."""
    match = re.match(r'=HYPERLINK\([^,]+,\s*"?([^"]+)"?\)', value, re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return value


def _clean_cell_value(value) -> str | None:
    """Normalize a cell value to a clean string, or None if empty."""
    if value is None:
        return None

    if isinstance(value, bool):
        return str(value)

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, int | float):
        return str(value)

    text = str(value).strip()
    if not text:
        return None

    # Handle Excel HYPERLINK formulas
    if text.upper().startswith("=HYPERLINK"):
        text = _extract_hyperlink_text(text)

    return text


def _match_header(header: str, candidates: list[str]) -> bool:
    """Check if a normalized header matches any candidate substring."""
    h = header.lower().replace("_", " ").strip()
    return any(c in h for c in candidates)


def _build_column_mapping(headers: list[str], tab_config: dict[str, list[str]]) -> dict[int, str]:
    """
    Map column indices to canonical field names based on tab-specific config.
    Returns {col_index: canonical_field_name}.
    """
    mapping: dict[int, str] = {}
    used_canonicals: set[str] = set()
    for idx, header in enumerate(headers):
        if header is None:
            continue
        for canonical, candidates in tab_config.items():
            if canonical in used_canonicals:
                continue
            if _match_header(header, candidates):
                mapping[idx] = canonical
                used_canonicals.add(canonical)
                break
    return mapping


class ExcelImporter:
    """Parses multi-tab CRM spreadsheets with per-tab column normalization."""

    def __init__(self, filepath: str | Path):
        self.filepath = Path(filepath)
        if not self.filepath.exists():
            raise FileNotFoundError(f"Excel file not found: {self.filepath}")

    def parse_all_tabs(self) -> dict[str, list[dict]]:
        """
        Parse all recognized tabs from the workbook.

        Returns:
            dict mapping tab name -> list of normalized row dicts.
        """
        wb = load_workbook(str(self.filepath), read_only=True, data_only=True)
        result: dict[str, list[dict]] = {}

        for sheet_name in wb.sheetnames:
            config_name = self._find_tab_config(sheet_name)
            if config_name is None:
                logger.info("Skipping unrecognized tab: %s", sheet_name)
                continue

            column_config = _TAB_COLUMN_MAPS[config_name]
            ws = wb[sheet_name]
            rows = self._parse_tab(ws, config_name, column_config)
            if rows:
                result[config_name] = rows
                logger.info(
                    "Parsed %d rows from tab '%s' (sheet '%s')",
                    len(rows),
                    config_name,
                    sheet_name,
                )

        wb.close()
        logger.info(
            "Import complete: %d tabs, %d total rows",
            len(result),
            sum(len(rows) for rows in result.values()),
        )
        return result

    def _find_tab_config(self, sheet_name: str) -> str | None:
        """Find matching tab configuration for a sheet name (exact match only)."""
        normalized = sheet_name.strip().lower()
        return _EXACT_SHEET_NAMES.get(normalized)

    def _parse_tab(
        self,
        ws,
        tab_name: str,
        column_config: dict[str, list[str]],
    ) -> list[dict]:
        """Parse a single worksheet tab into normalized row dicts."""
        rows_iter = ws.iter_rows()

        # Read header row
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return []

        headers = [_clean_cell_value(cell.value) for cell in header_row]
        col_mapping = _build_column_mapping(headers, column_config)

        if not col_mapping:
            logger.warning(
                "No columns matched for tab '%s'. Headers found: %s",
                tab_name,
                headers,
            )
            return []

        logger.debug(
            "Column mapping for '%s': %s",
            tab_name,
            {headers[idx]: name for idx, name in col_mapping.items()},
        )

        parsed_rows: list[dict] = []
        for row_num, row in enumerate(rows_iter, start=2):
            row_dict: dict[str, str | None] = {}
            has_data = False

            for col_idx, field_name in col_mapping.items():
                if col_idx < len(row):
                    value = _clean_cell_value(row[col_idx].value)
                    row_dict[field_name] = value
                    if value is not None:
                        has_data = True

            if not has_data:
                continue

            # Special handling: Family_Friends has names in column 0 with no header
            if tab_name == "Family_Friends_Mentors_Others" and len(row) > 0:
                name_val = _clean_cell_value(row[0].value)
                if name_val:
                    row_dict["name"] = name_val

            # Normalize: if we have "name" (full name) but no first/last, split it
            if row_dict.get("name") and not row_dict.get("first_name"):
                full_name = row_dict["name"]
                parts = full_name.split(None, 1)
                row_dict["first_name"] = parts[0] if parts else full_name
                row_dict["last_name"] = parts[1] if len(parts) > 1 else None

            # Skip rows with no company AND no email AND no name
            has_company = bool(row_dict.get("company") or row_dict.get("company_name"))
            has_email = bool(row_dict.get("email"))
            has_name = bool(row_dict.get("name") or row_dict.get("first_name"))
            if not has_company and not has_email and not has_name:
                continue

            row_dict["_source_row"] = str(row_num)
            parsed_rows.append(row_dict)

        return parsed_rows
